"""Excel出力クラス（テンプレート方式）

テンプレートのtempシートを複製→データ書込→保存後、
zipレベルでテンプレートのネイティブ図形を各シートのdrawingに注入する。
"""

import os
import re
import zipfile
from copy import deepcopy
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from lxml import etree
from openpyxl.writer.excel import ExcelWriter

from core.calc_utils import (
    parse_kilo, px_to_m_x, px_to_m_y,
    format_table_entry, circled_number,
)

EMU_PER_PX = 9525

# 画像ネイティブサイズ
IMG_W = 1274
IMG_H_FULL = 992
# トリミング: Y=900px以下をカット
IMG_H = 900
IMG_W_EMU = IMG_W * EMU_PER_PX
IMG_H_EMU = IMG_H * EMU_PER_PX

# テンプレートのデフォルトパス
DEFAULT_TEMPLATE_PATH = r"Q:\004_検測G\解析共用\UTRAS解析マクロ・マニュアル\アトラス報告書.xlsx"
TEMPLATE_SHEET_NAME = "temp"

# 変状一覧の配置
TABLE_START_ROW = 39
TABLE_ROWS_PER_COL = 4
TABLE_MAX_ENTRIES = 8

# 種別ごとの色 (RRGGBB)
CATEGORY_EXCEL_COLORS = {
    "ゆるみ": "0000FF",
    "空洞": "FF0000",
    "除外区間": "FF0000",
}


# テンプレートの画像表示サイズ（COMで取得したポイント値）
# Picture 3: Left=0, Top=0, Width=955.728pt, Height=669.293pt
_IMG_DISPLAY_W_PT = 955.728271484375
_IMG_DISPLAY_H_PT = 669.29345703125

# ピクセル→ポイント変換係数
_PX_TO_PT_X = _IMG_DISPLAY_W_PT / IMG_W   # 1274px → 955.73pt
_PX_TO_PT_Y = _IMG_DISPLAY_H_PT / IMG_H   # 900px → 669.29pt

# ポイント→EMU: 1pt = 12700 EMU
_PT_TO_EMU = 12700


def _px_to_emu(px):
    return int(px * EMU_PER_PX)


# テンプレートの列幅・行高さ（COMポイント値）
_COL_WIDTHS_PT = [411.75, 411.75, 7.5, 45.75, 79.5]
_ROW_HEIGHT_PT = 18.75

# 累積ポイント
_COL_CUM_PT = [0.0]
for _w in _COL_WIDTHS_PT:
    _COL_CUM_PT.append(_COL_CUM_PT[-1] + _w)


def _px_to_cell_anchor(px_x, px_y):
    """画像ピクセル座標をセルのAnchorMarkerに変換（COMポイント値ベース）"""
    # ピクセル→画像上のポイント位置
    pt_x = px_x * _PX_TO_PT_X
    pt_y = px_y * _PX_TO_PT_Y

    # X: 列を特定
    col = 0
    col_off_pt = pt_x
    for i in range(1, len(_COL_CUM_PT)):
        if pt_x < _COL_CUM_PT[i]:
            col = i - 1
            col_off_pt = pt_x - _COL_CUM_PT[i - 1]
            break
    else:
        col = len(_COL_WIDTHS_PT) - 1
        col_off_pt = pt_x - _COL_CUM_PT[-2]

    # Y: 行を特定
    row = int(pt_y / _ROW_HEIGHT_PT)
    row_off_pt = pt_y - row * _ROW_HEIGHT_PT

    return AnchorMarker(
        col=col, colOff=int(col_off_pt * _PT_TO_EMU),
        row=row, rowOff=int(row_off_pt * _PT_TO_EMU),
    )


class ExcelExporter:
    def __init__(self, image_groups, sorted_kilos, db):
        self.image_groups = image_groups
        self.sorted_kilos = sorted_kilos
        self.db = db

    def export(self, output_path, template_path=None):
        tpl = template_path or DEFAULT_TEMPLATE_PATH
        if not os.path.exists(tpl):
            raise FileNotFoundError(f"テンプレートが見つかりません: {tpl}")

        # --- Step 1: テンプレートからネイティブ図形XMLを抽出 ---
        tpl_native_anchors = self._extract_native_shapes(tpl)

        # --- Step 2: openpyxlでシート複製・データ書込・保存 ---
        wb = load_workbook(tpl)
        if TEMPLATE_SHEET_NAME not in wb.sheetnames:
            raise ValueError(f"テンプレートに '{TEMPLATE_SHEET_NAME}' シートがありません")

        temp_ws = wb[TEMPLATE_SHEET_NAME]

        # テンプレートのPNG画像を保持（copy_worksheetでコピーされないため）
        tpl_image_data = []
        for tpl_img in temp_ws._images:
            ref = tpl_img.ref
            if hasattr(ref, 'read'):
                ref.seek(0)
                img_bytes = ref.read()
            else:
                with open(ref, 'rb') as f:
                    img_bytes = f.read()
            tpl_image_data.append({
                'bytes': img_bytes,
                'width': tpl_img.width,
                'height': tpl_img.height,
                'anchor': tpl_img.anchor,
            })

        for kilo in self.sorted_kilos:
            ws = wb.copy_worksheet(temp_ws)
            ws.title = kilo

            # テンプレートのPNG画像を手動コピー
            for data in tpl_image_data:
                new_img = XlImage(BytesIO(data['bytes']))
                new_img.width = data['width']
                new_img.height = data['height']
                new_img.anchor = deepcopy(data['anchor'])
                ws.add_image(new_img)

            # 撮影画像配置（Y=900pxでトリミング、テンプレートと同じtwoCellAnchorで配置）
            group = self.image_groups[kilo]
            img_path = group.get('marked')
            if img_path:
                from PIL import Image as PILImage
                pil_img = PILImage.open(img_path)
                pil_img = pil_img.crop((0, 0, IMG_W, IMG_H))
                img_buf = BytesIO()
                pil_img.save(img_buf, format='JPEG', quality=95, dpi=(96, 96))
                img_buf.seek(0)
                img = XlImage(img_buf)
                img.width = IMG_W
                img.height = IMG_H
                # テンプレートの画像と同じtwoCellAnchor座標を使用
                img.anchor = TwoCellAnchor(editAs='oneCell')
                img.anchor._from = AnchorMarker(col=0, colOff=0, row=0, rowOff=0)
                img.anchor.to = AnchorMarker(col=4, colOff=1003024, row=35, rowOff=165652)
                ws.add_image(img)

            # フッター
            ws.oddFooter.center.text = "&P"

            # 変状シェイプ
            drawings = self.db.load_drawings(kilo)
            base_kilo_m = parse_kilo(kilo)
            direction = group.get('direction', '起点→終点')

            ws._custom_shape_trees = []
            shape_idx = 0
            for d in drawings:
                shape_idx += 1
                shape_tree = self._create_shape_tree(d, shape_idx)
                if shape_tree is not None:
                    ws._custom_shape_trees.append(shape_tree)

                category = d.get('category', 'ゆるみ')
                mgmt = d.get('mgmt_number')
                if mgmt and category != '除外区間':
                    label_tree = self._create_label_tree(d, mgmt, shape_idx)
                    if label_tree is not None:
                        ws._custom_shape_trees.append(label_tree)

            self._write_table(ws, drawings, base_kilo_m, direction)

        del wb[TEMPLATE_SHEET_NAME]
        self._save_with_shapes(wb, output_path)

        # --- Step 3: zip再処理でネイティブ図形を注入 ---
        if tpl_native_anchors:
            self._inject_native_shapes(output_path, tpl_native_anchors)

    # ------------------------------------------------------------------
    # テンプレートからネイティブ図形を抽出
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_native_shapes(tpl_path):
        """テンプレートのdrawing XMLからネイティブ図形（sp）のアンカーを抽出。
        画像(pic)アンカーは除外する。
        """
        anchors = []
        with zipfile.ZipFile(tpl_path, 'r') as z:
            for name in z.namelist():
                if re.match(r'xl/drawings/drawing\d+\.xml$', name):
                    tree = etree.fromstring(z.read(name))
                    for child in tree:
                        local = etree.QName(child.tag).localname
                        if 'anchor' not in local.lower():
                            continue
                        # pic要素がある＝画像アンカー → スキップ
                        has_pic = any(
                            'pic' == (etree.QName(e.tag).localname)
                            for e in child.iter()
                        )
                        if has_pic:
                            continue
                        # sp要素がある＝ネイティブ図形
                        has_sp = any(
                            'sp' == (etree.QName(e.tag).localname)
                            for e in child.iter()
                        )
                        if has_sp:
                            anchors.append(etree.tostring(child))
        return anchors

    # ------------------------------------------------------------------
    # zip再処理でネイティブ図形を注入
    # ------------------------------------------------------------------

    @staticmethod
    def _inject_native_shapes(xlsx_path, tpl_anchors):
        """保存済みxlsxの各シートのdrawing XMLにテンプレートのネイティブ図形を追加"""
        # 全エントリを読み込み
        entries = {}
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            for name in zin.namelist():
                entries[name] = zin.read(name)

        # 各drawing XMLにアンカーを追加
        for filename, data in list(entries.items()):
            if not re.match(r'xl/drawings/drawing\d+\.xml$', filename):
                continue
            tree = etree.fromstring(data)
            for anchor_bytes in tpl_anchors:
                anchor_elem = etree.fromstring(anchor_bytes)
                tree.append(anchor_elem)
            entries[filename] = etree.tostring(
                tree, xml_declaration=True, encoding='UTF-8', standalone=True,
            )

        # 書き出し
        tmp_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for filename, data in entries.items():
                zout.writestr(filename, data)
        os.replace(tmp_path, xlsx_path)

    # ------------------------------------------------------------------
    # 図形XML構築 (twoCellAnchor — COMポイント値ベースの正確な座標変換)
    # ------------------------------------------------------------------

    @staticmethod
    def _build_anchor_marker_xml(parent, tag, marker):
        XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        elem = etree.SubElement(parent, f"{{{XDR}}}{tag}")
        etree.SubElement(elem, f"{{{XDR}}}col").text = str(marker.col)
        etree.SubElement(elem, f"{{{XDR}}}colOff").text = str(marker.colOff)
        etree.SubElement(elem, f"{{{XDR}}}row").text = str(marker.row)
        etree.SubElement(elem, f"{{{XDR}}}rowOff").text = str(marker.rowOff)

    def _create_shape_tree(self, drawing, idx):
        XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        A = "http://schemas.openxmlformats.org/drawingml/2006/main"

        shape_type = drawing['type']
        category = drawing.get('category', 'ゆるみ')
        lx0, ly0 = drawing['lx0'], drawing['ly0']
        lx1, ly1 = drawing['lx1'], drawing['ly1']
        min_x, max_x = min(lx0, lx1), max(lx0, lx1)
        min_y, max_y = min(ly0, ly1), max(ly0, ly1)

        from_marker = _px_to_cell_anchor(min_x, min_y)
        to_marker = _px_to_cell_anchor(max_x, max_y)

        prst = 'rect' if (category == "除外区間" or shape_type == 'rectangle') else 'ellipse'
        color = CATEGORY_EXCEL_COLORS.get(category, '0000FF')

        anchor = etree.Element(f"{{{XDR}}}twoCellAnchor")
        self._build_anchor_marker_xml(anchor, "from", from_marker)
        self._build_anchor_marker_xml(anchor, "to", to_marker)

        sp = etree.SubElement(anchor, f"{{{XDR}}}sp")
        nvSpPr = etree.SubElement(sp, f"{{{XDR}}}nvSpPr")
        cNvPr = etree.SubElement(nvSpPr, f"{{{XDR}}}cNvPr")
        cNvPr.set("id", str(100 + idx))
        cNvPr.set("name", f"Shape {idx}")
        etree.SubElement(nvSpPr, f"{{{XDR}}}cNvSpPr")

        spPr = etree.SubElement(sp, f"{{{XDR}}}spPr")
        prstGeom = etree.SubElement(spPr, f"{{{A}}}prstGeom")
        prstGeom.set("prst", prst)
        etree.SubElement(prstGeom, f"{{{A}}}avLst")

        if category == "除外区間":
            sf = etree.SubElement(spPr, f"{{{A}}}solidFill")
            sc = etree.SubElement(sf, f"{{{A}}}srgbClr")
            sc.set("val", "FF0000")
            al = etree.SubElement(sc, f"{{{A}}}alpha")
            al.set("val", "30000")
        else:
            etree.SubElement(spPr, f"{{{A}}}noFill")

        ln = etree.SubElement(spPr, f"{{{A}}}ln")
        ln.set("w", "19050")
        lf = etree.SubElement(ln, f"{{{A}}}solidFill")
        lc = etree.SubElement(lf, f"{{{A}}}srgbClr")
        lc.set("val", color)

        etree.SubElement(anchor, f"{{{XDR}}}clientData")
        return anchor

    def _create_label_tree(self, drawing, mgmt_number, idx):
        XDR = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
        A = "http://schemas.openxmlformats.org/drawingml/2006/main"

        lx0, ly0 = drawing['lx0'], drawing['ly0']
        lx1, ly1 = drawing['lx1'], drawing['ly1']
        label_x = max(lx0, lx1) + 2
        label_y = max(0, min(ly0, ly1) - 2)

        num_text = circled_number(mgmt_number)
        category = drawing.get('category', 'ゆるみ')
        color = CATEGORY_EXCEL_COLORS.get(category, '0000FF')

        from_marker = _px_to_cell_anchor(label_x, label_y)
        to_marker = _px_to_cell_anchor(label_x + 22, label_y + 20)

        anchor = etree.Element(f"{{{XDR}}}twoCellAnchor")
        self._build_anchor_marker_xml(anchor, "from", from_marker)
        self._build_anchor_marker_xml(anchor, "to", to_marker)

        sp = etree.SubElement(anchor, f"{{{XDR}}}sp")
        nvSpPr = etree.SubElement(sp, f"{{{XDR}}}nvSpPr")
        cNvPr = etree.SubElement(nvSpPr, f"{{{XDR}}}cNvPr")
        cNvPr.set("id", str(200 + idx))
        cNvPr.set("name", f"Label {idx}")
        etree.SubElement(nvSpPr, f"{{{XDR}}}cNvSpPr")

        spPr = etree.SubElement(sp, f"{{{XDR}}}spPr")
        pg = etree.SubElement(spPr, f"{{{A}}}prstGeom")
        pg.set("prst", "rect")
        etree.SubElement(pg, f"{{{A}}}avLst")
        etree.SubElement(spPr, f"{{{A}}}noFill")
        ln = etree.SubElement(spPr, f"{{{A}}}ln")
        etree.SubElement(ln, f"{{{A}}}noFill")

        txBody = etree.SubElement(sp, f"{{{XDR}}}txBody")
        bp = etree.SubElement(txBody, f"{{{A}}}bodyPr")
        bp.set("wrap", "none")
        for attr in ("lIns", "tIns", "rIns", "bIns"):
            bp.set(attr, "0")
        p = etree.SubElement(txBody, f"{{{A}}}p")
        r = etree.SubElement(p, f"{{{A}}}r")
        rPr = etree.SubElement(r, f"{{{A}}}rPr")
        rPr.set("lang", "ja-JP")
        rPr.set("sz", "1800")
        rPr.set("b", "1")
        sf = etree.SubElement(rPr, f"{{{A}}}solidFill")
        sc = etree.SubElement(sf, f"{{{A}}}srgbClr")
        sc.set("val", color)
        t = etree.SubElement(r, f"{{{A}}}t")
        t.text = num_text

        etree.SubElement(anchor, f"{{{XDR}}}clientData")
        return anchor

    # ------------------------------------------------------------------
    # 変状一覧
    # ------------------------------------------------------------------

    def _write_table(self, ws, drawings, base_kilo_m, direction):
        table_drawings = [d for d in drawings if d.get('category') != '除外区間']
        table_drawings.sort(key=lambda d: d.get('mgmt_number') or 9999)
        table_drawings = table_drawings[:TABLE_MAX_ENTRIES]

        for i in range(TABLE_MAX_ENTRIES):
            col = 1 if i < TABLE_ROWS_PER_COL else 2
            row_idx = TABLE_START_ROW + (i % TABLE_ROWS_PER_COL)

            if i < len(table_drawings):
                d = table_drawings[i]
                mgmt = d.get('mgmt_number')
                area = d.get('area', '')
                category = d.get('category', 'ゆるみ')
                lx0, lx1 = d['lx0'], d['lx1']
                ly0, ly1 = d['ly0'], d['ly1']
                entry = format_table_entry(
                    mgmt, area, base_kilo_m, direction,
                    px_to_m_x(min(lx0, lx1)), px_to_m_x(max(lx0, lx1)),
                    px_to_m_y(min(ly0, ly1), area), px_to_m_y(max(ly0, ly1), area),
                    category,
                )
            else:
                entry = ""
            ws.cell(row=row_idx, column=col, value=entry)

    # ------------------------------------------------------------------
    # 保存（図形注入パッチ）
    # ------------------------------------------------------------------

    def _save_with_shapes(self, wb, output_path):
        _orig_write_ws = ExcelWriter.write_worksheet

        def _patched_write_ws(writer_self, ws):
            _orig_write_ws(writer_self, ws)
            shape_trees = getattr(ws, '_custom_shape_trees', [])
            if shape_trees and ws._drawing:
                drawing = ws._drawing
                _orig_sd_write = drawing._write

                def _patched_sd_write():
                    tree = _orig_sd_write()
                    for shape_tree in shape_trees:
                        tree.append(shape_tree)
                    return tree

                drawing._write = _patched_sd_write

        ExcelWriter.write_worksheet = _patched_write_ws
        try:
            wb.save(output_path)
        finally:
            ExcelWriter.write_worksheet = _orig_write_ws
