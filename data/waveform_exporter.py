"""連続波形画像のExcel出力（テンプレート方式）

連続表示している波形画像・ヒートマップを100m区切りでExcelに出力する。
テンプレートのtempシートを複製し、所定のセル座標に画像を配置する。
A～D列（合計4435px）、1枚につき2行（336px × 2）を使用して配置。
上部: 0～20m繰り返しスケール / 下部: キロ程スケール / 左: 深さスケール
"""

import os
import math
import zipfile
import re
from copy import deepcopy
from io import BytesIO
from PIL import Image as PILImage, ImageDraw, ImageFont
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from lxml import etree

from core.calc_utils import parse_kilo, px_to_m_x, m_to_px_y, WORK_AREAS

AREA_NAMES = ["左軌間外", "軌間内", "右軌間外"]
IMAGE_SPAN_M = 20.0

# 画像ピクセル範囲
X_PX_MIN = 72
X_PX_MAX = 916
STRIP_W = X_PX_MAX - X_PX_MIN   # 844px per 20m
PX_PER_M = STRIP_W / IMAGE_SPAN_M  # 42.2 px/m

# 各エリアの高さ（ピクセル）
AREA_STRIP_HEIGHTS = {
    name: area['y_max'] - area['y_min']
    for name, area in WORK_AREAS.items()
}

# 合成画像レイアウト定数
LABEL_TEXT_W = 100   # エリア名テキスト幅
DEPTH_SCALE_W = 55   # 深さスケール幅
LEFT_W = LABEL_TEXT_W + DEPTH_SCALE_W  # 合計左マージン
TOP_SCALE_H = 50     # 上部0-20mスケール高さ
SCALE_H = 55         # 下部キロ程スケール高さ
AREA_GAP = 16        # エリア間ギャップ
SEGMENT_M = 100      # 1行あたりの距離（m）
ROWS_PER_SHEET = 4   # 1シートあたりの行数（4行 × 100m = 400m）

# テンプレート設定
DEFAULT_TEMPLATE_PATH = r"Q:\004_検測G\解析共用\UTRAS解析マクロ・マニュアル\連続波形画像テンプレート.xlsx"
TEMPLATE_SHEET_NAME = "temp"

# 変状カテゴリ色（RGBA）
CATEGORY_COLORS_PIL = {
    "ゆるみ": (60, 120, 255, 100),
    "空洞": (255, 60, 60, 100),
}
CATEGORY_BORDER_PIL = {
    "ゆるみ": (60, 120, 255, 220),
    "空洞": (255, 60, 60, 220),
}
EXCLUSION_COLOR = (255, 100, 100, 80)

# 深さ目盛り値
DEPTH_TICKS = [0.0, 0.5, 1.0, 1.5, 2.0]

# フォント
_FONT_DIR = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')
_FONT_PATH = os.path.join(_FONT_DIR, 'meiryo.ttc')
_FONT_PATH_FALLBACK = os.path.join(_FONT_DIR, 'msgothic.ttc')


def _get_font(size):
    for path in (_FONT_PATH, _FONT_PATH_FALLBACK):
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


class WaveformExcelExporter:
    def __init__(self, image_groups, sorted_kilos, db):
        self.image_groups = image_groups
        self.sorted_kilos = sorted_kilos
        self.db = db

    def export(self, output_path, areas=None, overlays=None, image_key='unmarked', template_path=None):
        """Excel出力メイン処理"""
        if areas is None:
            areas = list(AREA_NAMES)
        if overlays is None:
            overlays = ["ゆるみ", "空洞", "除外区間"]

        if not self.sorted_kilos:
            raise ValueError("出力するデータがありません")

        tpl = template_path or DEFAULT_TEMPLATE_PATH
        if not os.path.exists(tpl):
            raise FileNotFoundError(f"テンプレートが見つかりません: {tpl}")

        # --- Step 1: テンプレートからネイティブ図形XMLを抽出 ---
        # テンプレートにロゴや専用シェイプが含まれている場合、openpyxlで消えないように退避
        tpl_native_anchors = self._extract_native_shapes(tpl)

        # 距離計算
        all_base_m = [parse_kilo(k) for k in self.sorted_kilos]
        data_start = min(all_base_m)
        data_end = max(all_base_m) + IMAGE_SPAN_M

        # 100m境界にアライン
        seg_start = math.floor(data_start / SEGMENT_M) * SEGMENT_M
        seg_end = math.ceil(data_end / SEGMENT_M) * SEGMENT_M

        segments_100m = []
        m = seg_start
        while m < seg_end:
            segments_100m.append(m)
            m += SEGMENT_M

        num_sheets = math.ceil(len(segments_100m) / ROWS_PER_SHEET)

        # フォント準備
        label_font = _get_font(22)
        scale_font = _get_font(20)
        depth_font = _get_font(16)
        top_scale_font = _get_font(18)

        # --- Step 2: openpyxlでシート複製・データ書込・保存 ---
        wb = load_workbook(tpl)
        temp_ws_name = TEMPLATE_SHEET_NAME if TEMPLATE_SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
        temp_ws = wb[temp_ws_name]

        # テンプレートに配置されている画像(ロゴ等)の保持
        tpl_image_data = []
        for tpl_img in temp_ws._images:
            ref = tpl_img.ref
            if hasattr(ref, 'read'):
                ref.seek(0)
                img_bytes = ref.read()
            else:
                with open(ref, 'rb') as f:
                    img_bytes = f.read()
            tpl_image_data.append({
                'bytes': img_bytes,
                'width': tpl_img.width,
                'height': tpl_img.height,
                'anchor': deepcopy(tpl_img.anchor),
            })

        for sheet_idx in range(num_sheets):
            start_idx = sheet_idx * ROWS_PER_SHEET
            end_idx = min(start_idx + ROWS_PER_SHEET, len(segments_100m))

            sheet_start = segments_100m[start_idx]
            sheet_end = segments_100m[end_idx - 1] + SEGMENT_M
            title = f"{self._fmt_kilo(sheet_start)}～{self._fmt_kilo(sheet_end)}"
            
            ws = wb.copy_worksheet(temp_ws)
            ws.title = title[:31]

            
            # テンプレートの画像を復元
            for data in tpl_image_data:
                new_img = XlImage(BytesIO(data['bytes']))
                new_img.width = data['width']
                new_img.height = data['height']
                new_img.anchor = deepcopy(data['anchor'])
                ws.add_image(new_img)

            # 1. テンプレート(temp_ws)から左ヘッダーやフッターを明示的にコピーして保護する
            ws.oddHeader.left.text = temp_ws.oddHeader.left.text

            # 各ブロックの画像を連続配置
            for block_idx in range(end_idx - start_idx):
                seg_m = segments_100m[start_idx + block_idx]

                composite = self._render_segment(
                    seg_m, seg_m + SEGMENT_M,
                    areas, overlays, image_key,
                    label_font, scale_font, depth_font, top_scale_font,
                )

                buf = BytesIO()
                # 印刷時の見切れ防止にDPI 96を指定
                composite.save(buf, format='PNG', optimize=True, dpi=(96, 96))
                buf.seek(0)

                xl_img = XlImage(buf)
                
                # 💡 テンプレート仕様に基づく配置 (1枚あたり2行使用)
                # エクセル上の行・列は0始まり
                # 画像1: row=0〜2(第1・2行), 画像2: row=2〜4(第3・4行) ...
                start_row = block_idx * 2
                end_row = start_row + 2
                
                # TwoCellAnchor を使って A〜D列(index 0〜3 -> end=4)、指定行に完全にフィットさせる
                anchor = TwoCellAnchor(editAs='oneCell')
                anchor._from = AnchorMarker(col=0, colOff=0, row=start_row, rowOff=0)
                anchor.to = AnchorMarker(col=4, colOff=0, row=end_row, rowOff=0)
                
                xl_img.anchor = anchor
                xl_img.width = composite.width
                xl_img.height = composite.height
                
                ws.add_image(xl_img)
                

        # テンプレート用シートを削除
        if temp_ws_name in wb.sheetnames:
            del wb[temp_ws_name]

        wb.save(output_path)

        # --- Step 3: zip再処理でネイティブ図形を注入 ---
        if tpl_native_anchors:
            self._inject_native_shapes(output_path, tpl_native_anchors)

    # ------------------------------------------------------------------
    # テンプレートからのネイティブ図形抽出・注入ロジック
    # ------------------------------------------------------------------

    @staticmethod
    def _extract_native_shapes(tpl_path):
        anchors = []
        with zipfile.ZipFile(tpl_path, 'r') as z:
            for name in z.namelist():
                if re.match(r'xl/drawings/drawing\d+\.xml$', name):
                    tree = etree.fromstring(z.read(name))
                    for child in tree:
                        local = etree.QName(child.tag).localname
                        if 'anchor' not in local.lower():
                            continue
                        has_pic = any('pic' == (etree.QName(e.tag).localname) for e in child.iter())
                        if has_pic:
                            continue
                        has_sp = any('sp' == (etree.QName(e.tag).localname) for e in child.iter())
                        if has_sp:
                            anchors.append(etree.tostring(child))
        return anchors

    @staticmethod
    def _inject_native_shapes(xlsx_path, tpl_anchors):
        entries = {}
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            for name in zin.namelist():
                entries[name] = zin.read(name)

        for filename, data in list(entries.items()):
            if not re.match(r'xl/drawings/drawing\d+\.xml$', filename):
                continue
            tree = etree.fromstring(data)
            for anchor_bytes in tpl_anchors:
                anchor_elem = etree.fromstring(anchor_bytes)
                tree.append(anchor_elem)
            entries[filename] = etree.tostring(
                tree, xml_declaration=True, encoding='UTF-8', standalone=True,
            )

        tmp_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for filename, data in entries.items():
                zout.writestr(filename, data)
        os.replace(tmp_path, xlsx_path)

    # ------------------------------------------------------------------
    # 100mセグメント合成画像の生成
    # ------------------------------------------------------------------

    def _render_segment(self, start_m, end_m, areas, overlays, image_key,
                        label_font, scale_font, depth_font, top_scale_font):
        seg_width_px = int((end_m - start_m) * PX_PER_M)
        total_area_h = (
            sum(AREA_STRIP_HEIGHTS[a] for a in areas)
            + AREA_GAP * max(0, len(areas) - 1)
        )
        # 右端の目盛り見切れ防止マージン
        RIGHT_MARGIN = 60
        
        comp_w = LEFT_W + seg_width_px + RIGHT_MARGIN
        comp_h = TOP_SCALE_H + total_area_h + SCALE_H

        composite = PILImage.new('RGB', (comp_w, comp_h), (30, 30, 30))
        draw = ImageDraw.Draw(composite)

        # 上部スケール
        self._draw_top_scale(draw, comp_w, start_m, end_m, top_scale_font)

        y_offset = TOP_SCALE_H
        for area_name in areas:
            area_def = WORK_AREAS[area_name]
            strip_h = AREA_STRIP_HEIGHTS[area_name]

            bbox = label_font.getbbox(area_name)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            tx = max(2, LABEL_TEXT_W - tw - 6)
            ty = y_offset + (strip_h - th) // 2
            draw.text((tx, ty), area_name, fill=(255, 255, 255), font=label_font)

            self._draw_depth_scale(draw, area_name, y_offset, strip_h, depth_font)

            draw.rectangle(
                [LEFT_W, y_offset, comp_w - 1, y_offset + strip_h - 1],
                fill=(45, 45, 50),
            )

            for kilo in self.sorted_kilos:
                base_m = parse_kilo(kilo)
                group = self.image_groups.get(kilo, {})
                direction = group.get('direction', '起点→終点')
                img_path = group.get(image_key)
                if not img_path or not os.path.exists(img_path):
                    continue
                if base_m + IMAGE_SPAN_M <= start_m or base_m >= end_m:
                    continue

                try:
                    pil_img = PILImage.open(img_path)
                    strip = pil_img.crop((
                        area_def['x_min'], area_def['y_min'],
                        area_def['x_max'], area_def['y_max'],
                    ))
                    if direction == '終点→起点':
                        strip = strip.transpose(PILImage.Transpose.FLIP_LEFT_RIGHT)

                    paste_x = LEFT_W + int((base_m - start_m) * PX_PER_M)
                    composite.paste(strip, (paste_x, y_offset))
                    pil_img.close()
                except Exception:
                    pass

            if overlays:
                self._draw_defect_overlay(
                    composite, area_name, y_offset, strip_h,
                    start_m, end_m, overlays,
                )

            # 20m毎の白色区切り線（目盛り線より太め）
            for kilo in self.sorted_kilos:
                base_m = parse_kilo(kilo)
                if base_m + IMAGE_SPAN_M <= start_m or base_m >= end_m:
                    continue
                sep_x = LEFT_W + int((base_m + IMAGE_SPAN_M - start_m) * PX_PER_M)
                if LEFT_W < sep_x < LEFT_W + seg_width_px:
                    draw.line(
                        [(sep_x, y_offset), (sep_x, y_offset + strip_h - 1)],
                        fill=(255, 255, 255), width=2,
                    )

            y_offset += strip_h + AREA_GAP

        # 下部スケール
        scale_y = comp_h - SCALE_H
        self._draw_bottom_scale(draw, scale_y, comp_w, start_m, end_m, scale_font)

        return composite

    # ------------------------------------------------------------------
    # スケール・オーバーレイ・ユーティリティ (変更なし)
    # ------------------------------------------------------------------

    def _draw_top_scale(self, draw, comp_w, start_m, end_m, font):
        draw.rectangle([LEFT_W, 0, comp_w - 1, TOP_SCALE_H - 1], fill=(25, 25, 30))
        draw.line([(LEFT_W, TOP_SCALE_H - 1), (comp_w - 1, TOP_SCALE_H - 1)], fill=(120, 120, 120), width=1)

        for kilo in self.sorted_kilos:
            base_m = parse_kilo(kilo)
            if base_m + IMAGE_SPAN_M <= start_m or base_m >= end_m:
                continue
            base_x = LEFT_W + int((base_m - start_m) * PX_PER_M)

            for rel_m in range(0, 21, 2):
                x = base_x + int(rel_m * PX_PER_M)
                if x < LEFT_W or x >= comp_w:
                    continue

                is_major = (rel_m % 10 == 0)
                is_mid = (rel_m % 5 == 0) and not is_major
                if is_major:
                    tick_h, color = 14, (220, 220, 220)
                elif is_mid:
                    tick_h, color = 10, (180, 180, 180)
                else:
                    tick_h, color = 6, (100, 100, 100)

                draw.line([(x, TOP_SCALE_H - 1 - tick_h), (x, TOP_SCALE_H - 1)], fill=color, width=1)

                if rel_m % 5 == 0:
                    label = f"{rel_m}m"
                    bbox = font.getbbox(label)
                    tw = bbox[2] - bbox[0]
                    if rel_m == 0:
                        lx = x
                    elif rel_m == 20:
                        lx = x - tw
                    else:
                        lx = x - tw // 2
                    draw.text((lx, 3), label, fill=(220, 220, 220), font=font)

            end_x = base_x + int(IMAGE_SPAN_M * PX_PER_M)
            if LEFT_W <= end_x < comp_w:
                draw.line([(end_x, 0), (end_x, TOP_SCALE_H - 1)], fill=(100, 100, 100), width=1)

    def _draw_depth_scale(self, draw, area_name, y_offset, strip_h, font):
        area_def = WORK_AREAS[area_name]
        y_min = area_def['y_min']

        draw.rectangle([LABEL_TEXT_W, y_offset, LEFT_W - 1, y_offset + strip_h - 1], fill=(35, 35, 40))

        for depth in DEPTH_TICKS:
            py = m_to_px_y(depth, area_name)
            local_y = int(py - y_min)
            abs_y = y_offset + local_y

            draw.line([(LEFT_W - 10, abs_y), (LEFT_W - 1, abs_y)], fill=(200, 200, 200), width=1)
            label = f"{depth:.1f}"
            bbox = font.getbbox(label)
            tw = bbox[2] - bbox[0]
            th = bbox[3] - bbox[1]
            draw.text((LEFT_W - 14 - tw, abs_y - th // 2), label, fill=(200, 200, 200), font=font)

        draw.line([(LEFT_W - 1, y_offset), (LEFT_W - 1, y_offset + strip_h - 1)], fill=(120, 120, 120), width=1)

    def _draw_defect_overlay(self, img, area_name, y_top, strip_h, start_m, end_m, overlays):
        overlay = PILImage.new('RGBA', img.size, (0, 0, 0, 0))
        ov_draw = ImageDraw.Draw(overlay)
        has_content = False

        for kilo in self.sorted_kilos:
            base_m = parse_kilo(kilo)
            if base_m + IMAGE_SPAN_M <= start_m or base_m >= end_m:
                continue

            group = self.image_groups.get(kilo, {})
            direction = group.get('direction', '起点→終点')
            drawings = self.db.load_drawings(kilo)

            for d in drawings:
                if d.get('area') != area_name:
                    continue
                category = d.get('category', 'ゆるみ')
                if category not in overlays:
                    continue

                lx0, lx1 = d['lx0'], d['lx1']
                min_lx, max_lx = min(lx0, lx1), max(lx0, lx1)
                offset_start = px_to_m_x(min_lx)
                offset_end = px_to_m_x(max_lx)

                if direction == "起点→終点":
                    abs_start = base_m + offset_start
                    abs_end = base_m + offset_end
                else:
                    abs_start = base_m - IMAGE_SPAN_M + offset_start
                    abs_end = base_m - IMAGE_SPAN_M + offset_end

                abs_start, abs_end = min(abs_start, abs_end), max(abs_start, abs_end)

                x0 = LEFT_W + int((abs_start - start_m) * PX_PER_M)
                x1 = LEFT_W + int((abs_end - start_m) * PX_PER_M)
                x0 = max(LEFT_W, x0)
                x1 = min(img.width, x1)
                if x0 >= x1:
                    continue

                has_content = True
                if category == '除外区間':
                    ov_draw.rectangle([x0, y_top, x1, y_top + 22], fill=EXCLUSION_COLOR, outline=(255, 80, 80, 180))
                else:
                    fill = CATEGORY_COLORS_PIL.get(category, (200, 200, 200, 80))
                    border = CATEGORY_BORDER_PIL.get(category, (200, 200, 200, 220))
                    ov_draw.rectangle([x0, y_top, x1, y_top + strip_h - 1], fill=fill, outline=border, width=2)

        if has_content:
            img_rgba = img.convert('RGBA')
            img_rgba = PILImage.alpha_composite(img_rgba, overlay)
            img.paste(img_rgba.convert('RGB'))

    def _draw_bottom_scale(self, draw, y, comp_w, start_m, end_m, font):
        draw.rectangle([LEFT_W, y, comp_w - 1, y + SCALE_H - 1], fill=(25, 25, 30))
        draw.line([(LEFT_W, y), (comp_w - 1, y)], fill=(150, 150, 150), width=1)

        step = 10
        m = start_m
        while m <= end_m:
            x = LEFT_W + int((m - start_m) * PX_PER_M)
            if x < LEFT_W or x >= comp_w:
                m += step
                continue

            is_major = (m % 100 == 0)
            is_mid = (m % 50 == 0) and not is_major
            if is_major:
                tick_h, color = 16, (240, 240, 240)
            elif is_mid:
                tick_h, color = 12, (200, 200, 200)
            elif m % 20 == 0:
                tick_h, color = 9, (160, 160, 160)
            else:
                tick_h, color = 5, (100, 100, 100)

            draw.line([(x, y), (x, y + tick_h)], fill=color, width=1)

            if m % 20 == 0:
                label = self._fmt_kilo(m)
                bbox = font.getbbox(label)
                tw = bbox[2] - bbox[0]
                lx = x - tw // 2
                draw.text((lx, y + 20), label, fill=(220, 220, 220), font=font)

            m += step

    @staticmethod
    def _fmt_kilo(m):
        km = int(m) // 1000
        rem = m % 1000
        return f"{km}k{rem:05.1f}m"